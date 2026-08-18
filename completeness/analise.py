# -*- coding: utf-8 -*-
"""Regras de negócio da auditoria: as 14 flags de completeness e a elegibilidade
dos SKUs pai (disponibilidade no site).

Nenhuma regra aqui pode mudar sem quebrar a comparabilidade da série histórica.
"""
from __future__ import annotations

import html
import os
import re
import unicodedata

import pandas as pd

import config


# --- Texto ------------------------------------------------------------------
def strip_html(s):
    if not s:
        return ''
    s = re.sub(r'<[^>]+>', ' ', s)
    return re.sub(r'\s+', ' ', html.unescape(s)).strip()


def norm(s):
    return unicodedata.normalize('NFKD', s or '').encode('ascii', 'ignore').decode().upper()


def sku_pai_de(sku: str) -> str:
    """SKU pai = SKU sem o último segmento, quando há mais de 3 pontos."""
    return '.'.join(sku.split('.')[:-1]) if sku.count('.') > 2 else sku


# --- Fontes auxiliares ------------------------------------------------------
def carregar_excecoes() -> set[str]:
    """SKUs desabilitados de propósito (planilha mantida pelo time de Produto)."""
    import openpyxl
    pasta = config.pasta_produto()
    if not pasta:
        print('AVISO: planilha de SKUs exceção não encontrada — nenhuma exceção aplicada.')
        return set()
    wb = openpyxl.load_workbook(os.path.join(pasta, config.ARQ_EXCECOES), read_only=True)
    it = wb.worksheets[0].iter_rows(values_only=True)
    next(it, None)
    exc = {str(r[1]).strip() for r in it if r[1]}
    wb.close()
    return exc


def carregar_fotos() -> dict[str, dict]:
    """SKUs (produto+cor) com foto de e-commerce nas pastas de coleção do Marketing.

    Nomes tipo 01.01.06.475.004_01.jpg ou 01.01.02.868.004.00_1.jpg -> produto+cor.
    Retorna {sku: {'rel': caminho p/ exibição, 'abs': caminho absoluto}}; o
    primeiro arquivo encontrado por SKU vence (ordem alfabética, coleção a coleção).
    """
    pastas = config.pastas_fotos()
    if not pastas:
        print('AVISO: pasta de fotos do Marketing não encontrada — contraprova de fotos não aplicada.')
        return {}
    varias = len(pastas) > 1
    skus: dict[str, dict] = {}
    for pasta in pastas:
        nome_col = os.path.basename(pasta.rstrip(os.sep))
        for root, _, files in os.walk(pasta):
            for f in sorted(files):
                m = re.match(r'^(\d{2}(?:\.\d{2,4})+)', f)
                if not m:
                    continue
                parts = m.group(1).split('.')
                if parts and parts[-1] == '00':
                    parts = parts[:-1]
                if len(parts) < 4:
                    continue
                sku = '.'.join(parts)
                if sku in skus:
                    continue
                caminho = os.path.join(root, f)
                rel = os.path.relpath(caminho, pasta).replace(os.sep, '/')
                # Com mais de uma coleção sincronizada, o caminho relativo sozinho
                # é ambíguo: prefixa a coleção para o time saber onde procurar.
                skus[sku] = {'rel': f'{nome_col}/{rel}' if varias else rel, 'abs': caminho}
    return skus


def carregar_base_produtos() -> tuple[dict, dict, dict, set]:
    """Uma passada na Base_Produtos, devolvendo os quatro mapas usados adiante.

    m         -> sku (pai ou filho) -> (coleção, grupo, status), p/ enriquecer o site
    sku_map   -> sk_produto -> (sku pai, tamanho), p/ agregar o estoque
    pai_attr  -> sku pai -> atributos de cadastro (item, cor, preço, grade...)
    com_envio -> SKUs com dt_envio preenchida (só estes podem estar ativos)
    """
    import openpyxl
    caminho = os.path.join(config.pasta_dados(), 'Base_Produtos.xlsx')
    wb = openpyxl.load_workbook(caminho, read_only=True)
    ws = wb['Consulta1']
    it = ws.iter_rows(values_only=True)
    hdr = list(next(it))
    cols = ['sk_produto', 'cod_produto', 'cod_sku_pai', 'desc_item', 'desc_colecao', 'desc_linha',
            'desc_grupo_wgb', 'desc_sub_grupo_wbg', 'desc_cor', 'desc_tamanho',
            'desc_status_produto', 'preco_tabela', 'dt_envio']
    ix = {c: hdr.index(c) for c in cols}
    ix_url = hdr.index('url') if 'url' in hdr else None   # foto do cadastro (Dim Produtos)
    m, sku_map, pai_attr, com_envio = {}, {}, {}, set()
    for r in it:
        pai = r[ix['cod_sku_pai']]
        if r[ix['dt_envio']] not in (None, 0, '', '0'):
            for k in (pai, r[ix['cod_produto']]):
                if k:
                    com_envio.add(k)
        val = (r[ix['desc_colecao']], r[ix['desc_grupo_wgb']], r[ix['desc_status_produto']])
        for c in (ix['cod_sku_pai'], ix['cod_produto']):
            k = r[c]
            if k and k not in m:
                m[k] = val
        if not pai:
            continue
        sku_map[r[ix['sk_produto']]] = (pai, r[ix['desc_tamanho']] or 'U')
        a = pai_attr.setdefault(pai, {'item': None, 'colecao': None, 'linha': None, 'grupo': None,
                                      'subgrupo': None, 'cor': None, 'status': None, 'preco': None,
                                      'dt_envio': None, 'url': None, 'tamanhos': set()})
        for key, col in [('item', 'desc_item'), ('colecao', 'desc_colecao'), ('linha', 'desc_linha'),
                         ('grupo', 'desc_grupo_wgb'), ('subgrupo', 'desc_sub_grupo_wbg'),
                         ('cor', 'desc_cor'), ('status', 'desc_status_produto'),
                         ('preco', 'preco_tabela'), ('dt_envio', 'dt_envio')]:
            if a[key] in (None, '') and r[ix[col]] not in (None, ''):
                a[key] = r[ix[col]]
        if ix_url is not None and a['url'] in (None, '') and r[ix_url] not in (None, ''):
            a['url'] = r[ix_url]
        a['tamanhos'].add(r[ix['desc_tamanho']] or 'U')
    wb.close()
    return m, sku_map, pai_attr, com_envio


# --- Catálogo do site -------------------------------------------------------
def analisar(prods: list[dict]) -> pd.DataFrame:
    """Aplica as 14 flags de completeness a cada produto publicado no site."""
    rows = []
    for pr in prods:
        title = pr['title'].strip()
        desc = strip_html(pr.get('body_html'))
        d = norm(desc)
        pt = (pr.get('product_type') or '').strip()
        first = norm(title).split()[0] if title else ''
        sub = config.SUB.get(first, 'Outros')
        if sub in config.ROUPA_SUBS:
            dept = 'Roupas'
        elif sub in config.ACESS_SUBS:
            dept = 'Acessórios'
        elif sub in config.CASA_SUBS:
            dept = 'Casa'
        else:
            if re.search(r'ROUPA|MALHA|TRICOT|JEANS|TECIDO PLANO(?!.*ACESS)', norm(pt)):
                dept = 'Roupas'
            elif re.search(r'BIJOUX|LENCO|CHAPEU|BOLSA|NECESSAIRE|CARTEIRA|CINTO|ACESS', norm(pt)):
                dept = 'Acessórios'
            elif re.search(r'DECORA|MESA|PAPELARIA|SAZONAL|COMPLEMENTAR', norm(pt)):
                dept = 'Casa'
            else:
                dept = 'Outros'
        variants = pr.get('variants') or []
        images = pr.get('images') or []
        price = min((float(v['price']) for v in variants if v.get('price')), default=0)
        grams = max((v.get('grams') or 0 for v in variants), default=0)
        skus = [1 for v in variants if (v.get('sku') or '').strip()]
        alts = [1 for i in images if (i.get('alt') or '').strip()
                and not re.fullmatch(r'image ?\d*', (i.get('alt') or '').strip(), re.I)]
        sku0 = next(((v.get('sku') or '').strip() for v in variants if (v.get('sku') or '').strip()), '')
        sku_pai = sku_pai_de(sku0)
        # Combined listing (pai que agrupa as cores): opção Cor/Color com >1 valor.
        # É duplicado interno — sai do dashboard, mas seus SKUs contam como "no site".
        combined = 1 if any((o.get('name') or '').lower() in ('cor', 'color') and
                            len([x for x in (o.get('values') or []) if x]) > 1
                            for o in (pr.get('options') or [])) else 0
        pais_all = sorted({sku_pai_de(s)
                           for s in ((v.get('sku') or '').strip() for v in variants) if s})
        f = [
            1 if len(desc) >= 30 else 0,
            1 if 'DETALHES DO PRODUTO' in d else 0,
            1 if re.search(r'COMPOSI|VISCOSE|ALGOD|LINHO|POLIESTER|POLIAMIDA|ELASTANO|SEDA|COURO|'
                           r'PALHA|CERAMIC|PORCELANA|MADEIRA|VIDRO|METAL|RESINA|LATAO|RATTAN|BAMBU|'
                           r'POLIPROPILENO|ACRILICO|LUREX|CASHMERE|LA\b', d) else 0,
            1 if re.search(r'\bCOR:', d) else 0,
            1 if re.search(r'MEDIDA|COMPRIMENTO|ALTURA|LARGURA|DIAMETRO|DIMENS|PROFUNDIDADE|'
                           r'\d+\s?CM\b|\d+\s?X\s?\d+', d) else 0,
            1 if 'OCASI' in d else 0,
            1 if len(images) >= 1 else 0,
            1 if len(images) >= 3 else 0,
            1 if images and len(alts) == len(images) else 0,
            1 if pt else 0,
            1 if len(pr.get('tags') or []) > 0 else 0,
            1 if variants and len(skus) == len(variants) else 0,
            1 if grams > 0 else 0,
            1 if price > 0 else 0,
        ]
        rows.append({'titulo': title, 'dept': dept, 'cat': sub, 'tipo_site': pt, 'sku_pai': sku_pai,
                     'combined': combined, 'pais_all': pais_all,
                     'preco': price, 'n_img': len(images), 'desc_len': len(desc),
                     'disponivel': 1 if any(v.get('available') for v in variants) else 0,
                     **{f'f{i}': f[i] for i in range(14)},
                     'url': f"https://www.souqstore.com.br/products/{pr['handle']}"})
    return pd.DataFrame(rows)


def enriquecer(df: pd.DataFrame, m: dict) -> pd.DataFrame:
    """Traz coleção/tipo/status do cadastro interno para o produto do site."""
    def look(s):
        for key in (s, '.'.join(s.split('.')[:-1])):
            if key in m:
                return m[key]
        return (None, None, None)

    res = df['sku_pai'].map(lambda s: look(s or ''))
    df['colecao'] = [r[0] or 'Sem cadastro' for r in res]
    df['tipo'] = [r[1] or 'Sem cadastro' for r in res]
    df['status'] = [r[2] or 'Sem cadastro' for r in res]
    return df


def imagens_do_site(prods: list[dict]) -> dict[str, str]:
    """sku pai -> 1ª imagem do Shopify (fallback de foto para o que está no site)."""
    img: dict[str, str] = {}
    for pr in prods:
        imagens = pr.get('images') or []
        if not imagens:
            continue
        src = imagens[0].get('src') or ''
        if not src:
            continue
        for v in (pr.get('variants') or []):
            s = (v.get('sku') or '').strip()
            if s:
                img.setdefault(sku_pai_de(s), src)
    return img


# --- Elegibilidade / disponibilidade ----------------------------------------
def disponibilidade(df_site, sku_map, pai_attr, com_envio, excecoes, fotos, imgs_site=None):
    """Candidatos = SKU pai com estoque >= 2 em localidade válida, fora das
    exceções e com data de envio, cruzados com a presença no site."""
    import openpyxl
    caminho = os.path.join(config.pasta_dados(), 'Base_Estoque.xlsx')
    wb = openpyxl.load_workbook(caminho, read_only=True)
    it = wb['Consulta1'].iter_rows(values_only=True)
    next(it)
    est = {}
    for r in it:
        sk, loc, q = r[0], r[2], r[3] or 0
        if loc in config.EXCL_LOCS or q <= 0:
            continue
        if sk not in sku_map:
            continue
        pai, tam = sku_map[sk]
        e = est.setdefault(pai, {'total': 0, 'cd': 0, 'lojas': set(), 'tam': set()})
        e['total'] += q
        if loc == config.CD_VENDAS:
            e['cd'] += q
        else:
            e['lojas'].add(loc)
        e['tam'].add(tam)
    wb.close()
    site_info = {}
    for _, r in df_site.iterrows():
        # todos os pais (produto+cor) do produto — inclui cores dentro de combined listings
        for pai in (r['pais_all'] or [r['sku_pai']]):
            site_info[pai] = max(site_info.get(pai, 0), r['disponivel'])
    rows = []
    n_exc, n_sem_envio = 0, 0
    for pai, e in est.items():
        if e['total'] < config.ESTOQUE_MIN_CANDIDATO:
            continue
        if pai in excecoes:
            n_exc += 1
            continue                      # SKU exceção: desabilitado de propósito
        if pai not in com_envio:
            n_sem_envio += 1
            continue                      # sem data de envio: não deveria estar ativo
        a = pai_attr.get(pai, {})
        grade_base = a.get('tamanhos') or e['tam']
        grade_pct = len(e['tam'] & grade_base) / len(grade_base) if grade_base else 1.0
        no_site = pai in site_info
        foto = fotos.get(pai) if hasattr(fotos, 'get') else None
        rows.append({'pai': pai, 'item': a.get('item') or pai,
                     'colecao': a.get('colecao') or 'Sem cadastro',
                     'linha': a.get('linha') or 'Sem cadastro',
                     'grupo': a.get('grupo') or 'Sem cadastro',
                     'subgrupo': a.get('subgrupo') or 'Sem cadastro',
                     'cor': a.get('cor') or '', 'status': a.get('status') or 'Sem cadastro',
                     'preco': a.get('preco') or 0,
                     'dt_envio': (a.get('dt_envio').strftime('%Y-%m-%d')
                                  if hasattr(a.get('dt_envio'), 'strftime')
                                  else (str(a.get('dt_envio'))
                                        if a.get('dt_envio') not in (None, 0, '', '0') else '')),
                     'qtde': e['total'], 'qtde_cd': e['cd'], 'n_filiais': len(e['lojas']),
                     'n_tam_estoque': len(e['tam']), 'n_tam_grade': len(grade_base),
                     'grade_pct': round(grade_pct, 3), 'grade_completa': 1 if grade_pct >= 1 else 0,
                     'no_site': 1 if no_site else 0,
                     'vendavel_site': 1 if site_info.get(pai) == 1 else 0,
                     'tem_foto': 1 if foto else 0,
                     'foto_arq': foto['rel'] if foto else '',
                     # Fallbacks de imagem para o app: cadastro (Dim Produtos) e
                     # 1ª foto do Shopify (só existe para quem já está no site).
                     'foto_url': a.get('url') or '',
                     'img_site': (imgs_site or {}).get(pai, '')})
    return pd.DataFrame(rows), n_exc, n_sem_envio
